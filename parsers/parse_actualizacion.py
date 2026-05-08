"""
parse_actualizacion.py
======================
Lee Peralta_I_Actualizacion_Rodamientos.xlsx y genera:
  - data/estado_actual.json    → estado y última medida por turbina
  - data/warnings.json         → warnings mensuales y por tipo por turbina

USO:
  python parsers/parse_actualizacion.py

REQUISITOS:
  pip install openpyxl
"""

import json
import os
import sys
from datetime import datetime
import openpyxl

# ── Rutas ──────────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_FILE = os.path.join(BASE_DIR, "Peralta_I_Actualizacion_Rodamientos.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "data")

# ── Helpers ───────────────────────────────────────────────────────────────────

def fmt_fecha(val):
    """Convierte datetime o string de fecha a 'YYYY-MM-DD', o None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    return s if s else None

def limpiar(val):
    """Devuelve None si el valor está vacío, sino el valor limpio."""
    if val is None:
        return None
    if isinstance(val, str):
        return val.strip() or None
    return val

# ── 1. Hoja Estado_Rodamientos → estado_actual.json ──────────────────────────

def parsear_estado(ws):
    """
    Estructura de la hoja (fila 4 = cabecera, datos desde fila 5):
      col 0  WEC
      col 1  Cat Del (última)
      col 2  Cat Tras (última)
      col 3  Fecha última muestra
      col 4  Tipo última muestra
      col 5  Cambio Rod. Frontal
      col 6  Cambio Rod. Trasero
      col 7  Insp. Joao (fecha)
      col 8  Insp. Joao (clase)
      col 9  Insp. Joao (comentario)
      col 10 Parada retrofit (días)
      col 11 Nueva fecha muestra      ← campos editables para próximo control
      col 12 Nuevo tipo muestra
      col 13 Nuevo Cat Del
      col 14 Nuevo Cat Tras
      col 15 Comentarios nuevos
    """
    resultado = {}

    for row in ws.iter_rows(min_row=5, values_only=True):
        wec = limpiar(row[0])
        if not wec:
            continue

        # Campos editables: si hay nueva medida, se usa; si no, la anterior
        nueva_fecha  = fmt_fecha(row[11])
        nueva_cat_del  = limpiar(row[13])
        nueva_cat_tras = limpiar(row[14])

        resultado[wec] = {
            # Estado actual (última medida confirmada)
            "cat_del":             limpiar(row[1]),
            "cat_tras":            limpiar(row[2]),
            "fecha_ultima":        fmt_fecha(row[3]),
            "tipo_ultima":         limpiar(row[4]),

            # Cambios de rodamiento
            "cambio_frontal":      fmt_fecha(row[5]),
            "cambio_trasero":      fmt_fecha(row[6]),

            # Inspección Joao Alvez
            "insp_joao_fecha":     fmt_fecha(row[7]),
            "insp_joao_clase":     limpiar(row[8]),
            "insp_joao_comentario": limpiar(row[9]),

            # Parada retrofit
            "parada_retrofit_dias": limpiar(row[10]),

            # Nueva medida (a completar en cada actualización)
            "nueva_fecha":         nueva_fecha,
            "nuevo_tipo":          limpiar(row[12]),
            "nueva_cat_del":       nueva_cat_del,
            "nueva_cat_tras":      nueva_cat_tras,
            "comentarios_nuevos":  limpiar(row[15]),
        }

        # Promover nueva medida a estado confirmado cuando está presente
        nuevo_tipo = limpiar(row[12])
        if nueva_cat_del:
            resultado[wec]["cat_del"] = nueva_cat_del
        if nueva_cat_tras:
            resultado[wec]["cat_tras"] = nueva_cat_tras
        if nueva_fecha:
            resultado[wec]["fecha_ultima"] = nueva_fecha
            if nuevo_tipo:
                resultado[wec]["tipo_ultima"] = nuevo_tipo

    return resultado


# ── 2. Hoja Nuevo_Control_De_Rodamientos → actualiza estado_actual.json ────────

def parsear_nuevo_control(ws):
    """
    Lee la hoja Nuevo_Control_De_Rodamientos y devuelve un dict con
    los datos nuevos por turbina.

    Estructura de la hoja (fila 1 = cabecera):
      col 0  Nueva fecha muestra
      col 1  Nuevo tipo muestra
      col 2  Nuevo Cat Del
      col 3  Nuevo Cat Tras
      col 4  Comentarios nuevos
      col 5+ PSP-01, PSP-02, ... PSP-50  (X = aplica a esa turbina)

    Cada fila es un nuevo control. Una turbina puede tener varios controles
    en distintas fechas — se queda con el más reciente.
    """

    # Leer headers para mapear columna → turbina
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row)

    # Construir mapa: nombre_turbina → índice de columna
    # Las turbinas empiezan en col 5 (PSP-01 = col 5, PSP-02 = col 6, etc.)
    turbina_cols = {}
    for i, h in enumerate(headers[5:], start=5):
        if h:
            # Normalizar PSP41 → PSP-41
            nombre = str(h).strip().replace('PSP', 'PSP-').replace('--', '-')
            turbina_cols[nombre] = i

    resultado = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Ignorar filas vacías
        if not any(row):
            continue

        nueva_fecha  = fmt_fecha(row[0])
        nuevo_tipo   = limpiar(row[1])
        nueva_cat_del  = limpiar(row[2])
        nueva_cat_tras = limpiar(row[3])
        comentarios  = limpiar(row[4])

        # Ignorar filas sin fecha (no completadas)
        if not nueva_fecha:
            continue

        # Aplicar a las turbinas marcadas con X
        for turbina, col_i in turbina_cols.items():
            if col_i < len(row) and limpiar(row[col_i]) in ('X', 'x'):
                # Si ya hay un control para esta turbina, quedarse con el más reciente
                if turbina not in resultado or nueva_fecha > resultado[turbina]['nueva_fecha']:
                    resultado[turbina] = {
                        "nueva_fecha":     nueva_fecha,
                        "nuevo_tipo":      nuevo_tipo,
                        "nueva_cat_del":   nueva_cat_del,
                        "nueva_cat_tras":  nueva_cat_tras,
                        "comentarios_nuevos": comentarios,
                    }

    return resultado


def aplicar_nuevo_control(estado, nuevo_control):
    """
    Combina el estado histórico con los datos del nuevo control.
    Si una turbina tiene datos nuevos, actualiza sus campos.
    Si no tiene, deja el estado como estaba.
    """
    for turbina, datos_nuevos in nuevo_control.items():
        if turbina not in estado:
            continue
        estado[turbina].update(datos_nuevos)
        # Promover a estado confirmado
        if datos_nuevos.get("nueva_cat_del"):
            estado[turbina]["cat_del"] = datos_nuevos["nueva_cat_del"]
        if datos_nuevos.get("nueva_cat_tras"):
            estado[turbina]["cat_tras"] = datos_nuevos["nueva_cat_tras"]
        if datos_nuevos.get("nueva_fecha"):
            estado[turbina]["fecha_ultima"] = datos_nuevos["nueva_fecha"]
            if datos_nuevos.get("nuevo_tipo"):
                estado[turbina]["tipo_ultima"] = datos_nuevos["nuevo_tipo"]

    return estado


# ── 3. Hoja Warnings_por_Tipo + Warnings_Mensuales → warnings.json ────────────

def parsear_warnings(ws_tipo, ws_mensual):
    """
    Warnings_por_Tipo (fila 4 = cabecera, datos desde fila 5):
      col 0  WEC
      col 1  Cat Tras
      col 2  Lub. Sin presión
      col 3  Lub. Sin pres.(aux)
      col 4  Grasa vacío
      col 5  Ruido Nacelle
      col 6  Sensor acúst.
      col 7  Air Gap
      col 8  TOTAL

    Warnings_Mensuales (fila 4 = cabecera, datos desde fila 5):
      col 0   WEC
      col 1   Cat Tras
      col 2…N meses (Ene'25, Feb'25, ...)
      col N+1 Nuevo mes (fecha)   ← campo editable
      col N+2 Nuevo mes (total)   ← campo editable
    """
    resultado = {}

    # --- Por tipo ---
    tipo_headers = []
    for row in ws_tipo.iter_rows(min_row=4, max_row=4, values_only=True):
        tipo_headers = [limpiar(c) for c in row]

    for row in ws_tipo.iter_rows(min_row=5, values_only=True):
        wec = limpiar(row[0])
        if not wec:
            continue
        resultado[wec] = {
            "cat_tras": limpiar(row[1]),
            "por_tipo": {},
            "mensual":  {},
            "nuevo_mes_fecha": None,
            "nuevo_mes_total": None,
        }
        for i, header in enumerate(tipo_headers[2:], start=2):
            if header and header != "TOTAL":
                resultado[wec]["por_tipo"][header] = limpiar(row[i]) or 0
        resultado[wec]["total_warnings"] = limpiar(row[8]) or 0

    # --- Mensuales ---
    mensual_headers = []
    for row in ws_mensual.iter_rows(min_row=4, max_row=4, values_only=True):
        mensual_headers = [limpiar(c) for c in row]

    # Identificar columnas de meses reales vs. campos editables
    # Los campos editables tienen salto de línea en el header
    meses_cols = []
    nuevo_mes_fecha_col = None
    nuevo_mes_total_col = None

    for i, h in enumerate(mensual_headers[2:], start=2):
        if h is None:
            continue
        if "Nuevo mes" in str(h) and "fecha" in str(h).lower():
            nuevo_mes_fecha_col = i
        elif "Nuevo mes" in str(h) and "total" in str(h).lower():
            nuevo_mes_total_col = i
        else:
            meses_cols.append((i, h))

    for row in ws_mensual.iter_rows(min_row=5, values_only=True):
        wec = limpiar(row[0])
        if not wec or wec not in resultado:
            continue
        for col_i, mes_nombre in meses_cols:
            if col_i < len(row):
                resultado[wec]["mensual"][mes_nombre] = limpiar(row[col_i]) or 0

        # Campos editables para el próximo mes
        if nuevo_mes_fecha_col and nuevo_mes_fecha_col < len(row):
            resultado[wec]["nuevo_mes_fecha"] = fmt_fecha(row[nuevo_mes_fecha_col])
        if nuevo_mes_total_col and nuevo_mes_total_col < len(row):
            resultado[wec]["nuevo_mes_total"] = limpiar(row[nuevo_mes_total_col])

    return resultado


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not os.path.exists(INPUT_FILE):
        print(f"ERROR: No se encontró '{INPUT_FILE}'")
        print("Asegurate de correr este script desde la carpeta del proyecto.")
        sys.exit(1)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print(f"Leyendo {INPUT_FILE}...")
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)

    # Estado histórico base
    ws_estado = wb["Estado_Rodamientos"]
    estado = parsear_estado(ws_estado)

    # Nuevos controles del mes — sobreescribe donde haya X
    ws_nuevo = wb["Nuevo_Control_De_Rodamientos"]
    nuevo_control = parsear_nuevo_control(ws_nuevo)

    if nuevo_control:
        print(f"  → Nuevos controles encontrados: {sorted(nuevo_control.keys())}")
        estado = aplicar_nuevo_control(estado, nuevo_control)
    else:
        print(f"  → Sin controles nuevos en 'Nuevo_Control_De_Rodamientos'")

    out_estado = os.path.join(OUTPUT_DIR, "estado_actual.json")
    with open(out_estado, "w", encoding="utf-8") as f:
        json.dump(estado, f, ensure_ascii=False, indent=2)
    print(f"  ✓ {out_estado}  ({len(estado)} turbinas)")

    # Warnings
    ws_tipo    = wb["Warnings_por_Tipo"]
    ws_mensual = wb["Warnings_Mensuales"]
    warnings = parsear_warnings(ws_tipo, ws_mensual)
    out_warnings = os.path.join(OUTPUT_DIR, "warnings.json")
    with open(out_warnings, "w", encoding="utf-8") as f:
        json.dump(warnings, f, ensure_ascii=False, indent=2)
    print(f"  ✓ {out_warnings}  ({len(warnings)} turbinas)")

    # Actualizar historico_rodamientos.json
    historico_path = os.path.join(OUTPUT_DIR, "historico_rodamientos.json")
    if os.path.exists(historico_path):
        with open(historico_path, 'r', encoding='utf-8') as f:
            historico = json.load(f)

        n_medidas      = 0
        n_inspecciones = 0

        for turbina, datos in estado.items():
            if turbina not in historico:
                continue

            hist = historico[turbina]

            # Actualizar campos de cambio de rodamiento
            for field in ("cambio_frontal", "cambio_trasero"):
                if datos.get(field) and hist.get(field) != datos[field]:
                    hist[field] = datos[field]

            # Actualizar inspección Joao Alvez
            if datos.get("insp_joao_fecha"):
                nueva_insp = {
                    "realizada": True,
                    "fecha": datos["insp_joao_fecha"],
                    "comentario": datos.get("insp_joao_comentario"),
                }
                if hist.get("insp_joao") != nueva_insp:
                    hist["insp_joao"] = nueva_insp
                    n_inspecciones += 1

            # Agregar nueva medida al historial si no existe
            nueva_fecha = datos.get("nueva_fecha")
            if nueva_fecha:
                fechas_existentes = {m["fecha"] for m in hist.get("medidas", [])}
                if nueva_fecha not in fechas_existentes:
                    hist.setdefault("medidas", []).append({
                        "fecha":    nueva_fecha,
                        "tipo":     datos.get("nuevo_tipo"),
                        "cat_del":  datos.get("nueva_cat_del"),
                        "cat_tras": datos.get("nueva_cat_tras"),
                    })
                    hist["medidas"].sort(key=lambda m: m.get("fecha") or "")
                    n_medidas += 1

        if n_medidas or n_inspecciones:
            with open(historico_path, 'w', encoding='utf-8') as f:
                json.dump(historico, f, ensure_ascii=False, indent=2)
            print(f"  ✓ {historico_path}  ({n_medidas} medidas nuevas, {n_inspecciones} inspecciones actualizadas)")
        else:
            print(f"  → {historico_path}  (sin cambios)")

    wb.close()
    print("\nListo. Los JSON están en la carpeta 'data/'.")


if __name__ == "__main__":
    main()
